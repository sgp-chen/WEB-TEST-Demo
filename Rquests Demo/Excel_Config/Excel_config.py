import openpyxl
from Tools.tools import Requests_tools
from my_config.log_config import get_log
log= get_log("../my_config/log.ini")

class Excel_Config():
    def Case_Run(self, i,log, path):
        excel = openpyxl.load_workbook(path)
        for sheets in excel.sheetnames:
            sheet = excel[sheets]
            for value in sheet.values:
                if type(value[0]) is int:
                    if value[2] is None:
                        pass
                    else:
                        url = value[2]
                    if value[5] is None:
                        pass
                    else:
                        code = value[5]
                    data = {}
                    data["headers"] = value[3]
                    data["data"] = value[4]
                    if value[6] == int(i):
                        log.info('正在执行：{}'.format(value[7]))
                        for key in list(data.keys()):
                            if data[key] is None:
                                del data[key]
                        if "assert" in value[1]:
                            re = Requests_tools()
                            getattr(re, value[1])(code, log,url, **data)
                        else:
                            re = Requests_tools()
                            getattr(re, value[1])(url, **data)
                    else:
                        pass
        excel.save(path)
        excel.close()

    def excel_value(self, path):
        excel = openpyxl.load_workbook(path)
        for sheets in excel.sheetnames:
            sheet = excel[sheets]
            rows = sheet.max_row
            value = sheet.cell(row=rows, column=7).value
        excel.close()
        return value


if __name__ == '__main__':
    Excel_Config().Case_Run(1, log,"../data/新建 XLSX 工作表.xlsx")
