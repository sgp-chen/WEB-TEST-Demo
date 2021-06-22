'''
    Excel数据驱动效果实现。
        实现目的：基于excel中的内容，来调用关键字类实现自动化测试的执行。
        相当于excel文件就是一个测试用例，底层代码就是关键字驱动类以及excel驱动类
'''
import logging.config

import openpyxl

# 获取到excel，进入sheet页中，读取单元格内容
from openpyxl.styles import PatternFill, Font

from excel_driver import excel_conf
from my_conf import log_conf
from ui_keys.web_keys import WebKey


# Excel数据驱动类
def excel_runner(path, log):
    excel = openpyxl.load_workbook(path)
    try:
        # 获取所有的sheet页
        sheets = excel.sheetnames
        # 遍历所有sheet页
        for sheet in sheets:
            sheet_temp = excel[sheet]
            # 如果sheet中包含特定字段，直接continue
            # 遍历sheet页中所有的单元格
            log.info('—————————{}—————————'.format(sheet))
            for values in sheet_temp.values:
                # 1. 读取用例的执行部分的内容。
                if type(values[0]) is int:
                    log.info('正在执行：{}'.format(values[5]))
                    # 提取本行的测试数据
                    data = {}
                    data['name'] = values[2]
                    data['value'] = values[3]
                    data['txt'] = values[4]
                    data['expect'] = values[6]
                    # print(data)
                    # 优化测试数据内容,将所有为None的数据全部清除出data中
                    for key in list(data.keys()):
                        # print(key)
                        if data[key] is None:
                            del data[key]
                    # 调用对应的关键字来执行操作行为：分为三种不同的场景，不同场景需要不同处理
                    if values[1] == 'open_browser':
                        wk = WebKey(values[4], log)
                    # 断言可能不会只有一种，只要有assert关键字，就是一个断言函数
                    elif 'assert' in values[1]:
                        # 只有断言函数才会有返回值。
                        status = getattr(wk, values[1])(**data)
                        # 基于status来写入测试的结果
                        if status:
                            # 执行Pass写入
                            excel_conf.pass_(sheet_temp.cell, values[0] + 2, 8)
                        else:
                            # 执行Failed写入
                            excel_conf.failed(sheet_temp.cell, values[0] + 2, 8)
                        # 执行excel的保存
                        excel.save(path)
                    else:
                        getattr(wk, values[1])(**data)
    except Exception as e:
        log.exception('运行异常：{}'.format(e))
    finally:
        # 关闭excel
        excel.close()


# 当这个类被其他类调用了，则该类下方的main函数不再生效
if __name__ == '__main__':
    print('这是excel读写的类')
