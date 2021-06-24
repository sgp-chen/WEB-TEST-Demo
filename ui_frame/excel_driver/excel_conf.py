'''
    Excel的配置文件
'''
import openpyxl
from openpyxl.styles import PatternFill, Font


# Pass配置
def pass_(cell, row, column):
    cell(row=row, column=column).value = 'Pass'
    # 写入单元格样式设定：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
    cell(row=row, column=column).font = Font(bold=True)


# Failed配置
def failed(cell, row, column):
    cell(row=row, column=column).value = 'Failed'
    # 写入单元格样式设定：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
    cell(row=row, column=column).font = Font(bold=True)
def excel_open(excel_path):
    excel=openpyxl.load_workbook(excel_path)
    return excel
def excel_sheet(path,sheet_name):
    sheets=excel_open(path)[sheet_name].values
    case_list = list()
    for value in sheets:
        if value[0] !="编号":
            case_list.append(value[1])
    excel_open(path).close()
    return case_list

if __name__ == '__main__':
    excel_sheet("../data/data.xlsx","Sheet2")
